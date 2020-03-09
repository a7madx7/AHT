"""
    CopyRight Dr. Ahmad Hamdi Emara 2020
    Adam Medical Company
    This module is intended for Order Evaluation and Color coding each item in the order.
"""
# =-=-=-=-=-=-=-=-=-=-=-=-=-=-==-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=
# IMPORTS REGION.
# =-=-=-=-=-=-=-=-=-=-=-=-=-=-==-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=
from openpyxl import Workbook, load_workbook

import jinja2
import pandas as pd
import numpy as np

import math
import datetime

from pathlib import Path
import os
# =-=-=-=-=-=-=-=-=-=-=-=-=-=-==-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=
# LOGIC REGION.
# =-=-=-=-=-=-=-=-=-=-=-=-=-=-==-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=
def evaluateOrder(file, df, pharmacy_name):
    try:
        export_file_path = get_export_file_path(file, pharmacy_name)

        # drop unnecessary columns
        drop_unnecessary_columns(df)

        remove_new_lines(df)
        remove_unnecessary_strings(df)

        # insert the MIN, MAX, RECOMMENDED COLUMNS
        insert_new_columns(df)
        seed_new_columns(df)

        # print(df.head())

        df.style.apply(highlight, axis=None).to_excel(export_file_path, sheet_name=get_file_id(pharmacy_name), na_rep='', float_format=None, columns=None, header=True, index=None, index_label=None, startrow=0, startcol=0, engine=None, merge_cells=True, encoding=None, inf_rep='inf', verbose=True, freeze_panes=None) 
        
        format_file(export_file_path)
        return True
    except Exception as e:
        print(e)
        return False

# =-=-=-=-=-=-=-=-=-=-=-=-=-=-==-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=
# IO REGION.
# =-=-=-=-=-=-=-=-=-=-=-=-=-=-==-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=

def get_file_id(pharmacy_name):
    # renew the now variable in case of leaving the app open for days.
    now = datetime.datetime.now()
    try:
        # return the file name with the pharmacy name inside.
        return f'{pharmacy_name} {now.strftime("%A")} {str(now.day)}-{str(now.month)}=={now.hour}-{now.minute}-{now.second}'
    except:
        # if there's any problem doing the above, return the file name without the pharmacy name.
        return f'Revised on {now.strftime("%A")} {str(now.day)}-{str(now.month)}=={now.hour}-{now.minute}-{now.second}'

def get_export_file_path(file, pharmacy_name):
    try:  
        _dir = Path(os.path.dirname(file)) ## directory of file
        export_file_path = _dir / str("REVISED ORDER " + get_file_id(pharmacy_name) + '.xlsx')
        return export_file_path
    except Exception as e:
        print(e)
        return e
# =-=-=-=-=-=-=-=-=-=-=-=-=-=-==-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=
# TRIMMING REGION.
# =-=-=-=-=-=-=-=-=-=-=-=-=-=-==-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=
def drop_unnecessary_columns(df):
    # drop unnecessary columns from the data frame.
    for i in range(1,21):
        index = f'None.{i}'
        del df[index]
    df.dropna(how='all', axis=1, inplace = True)

def remove_new_lines(df):
    for column in df:
        if column is not None:
            df.rename(columns={column: column.replace('\n',' ')}, inplace=True)
            column = column.replace('\n',' ')

def remove_unnecessary_strings(df):
    # Remove the unnecessary ## LOC.:DEFAULT
    df['PRODUCT'] = df['PRODUCT'].apply(lambda x: x.split('\n')[0])

# =-=-=-=-=-=-=-=-=-=-=-=-=-=-==-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=
# ALTERING/MODIFYING REGION.
# =-=-=-=-=-=-=-=-=-=-=-=-=-=-==-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=
def format_file(file):
    try:
        wb = load_workbook(file)
        ws = wb.worksheets[0]

        ws.column_dimensions['A'].width = 60
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['F'].width = 15
        ws.column_dimensions['G'].width = 10
        ws.column_dimensions['H'].width = 15
        ws.column_dimensions['J'].width = 15
        ws.column_dimensions['K'].width = 20

        ws.sheet_properties.tabColor = "F17CF7"
        wb.save(filename = file)
        return True
    except Exception as e:
        print(e)
        return False

def insert_new_columns(df):
    # insert the min column
    df.insert(2, 'MIN', 0)
    # insert the max column
    df.insert(2, 'MAX', 0)
    # insert the recommended column
    df.insert(2, 'RECOMMENDED', 0)

def seed_new_columns(df):
    # Initial seeding for those columns.
    df['MIN'] = df['AUTO ORDER'].apply(lambda x: math.ceil(0.5 * x))
    df['MAX'] = df['AUTO ORDER'].apply(lambda x: math.ceil(1.5 * x))
    df['RECOMMENDED'] = (df['AUTO ORDER'] + df['BRANCH STOCK']).apply(lambda x: math.ceil(x / 4.0))
    # =-=-=-=-=-=-=-=-=-=-=-=-=-=-==-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=
    # Fix false recommendation.
    false_reccomendation = df['AUTO ORDER'] <= 0
    df.loc[false_reccomendation, 'RECOMMENDED'] = 0

def highlight(x):
    # set the highlight style for each condition.
    approved_max_style = 'background-color: pink; font-size: 16pt; text-align: center'
    approved_style = 'background-color: #90ee90; font-size: 16pt; text-align: center;'
    rejected_min_style = 'background-color: #F7F57C; font-size: 16pt; text-align: center;'
    rejected_max_style = 'background-color: #F7857C; font-size: 16pt; text-align: center;'
    approved_addition_style = 'background-color: #7CBCF7; font-size: 16pt; text-align: center;'

    # =-=-=-=-=-=-=-=-=-=-=-=-=-=-==-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=
    # get a clone of the original data frame.
    df1 = pd.DataFrame('', index=x.index, columns=x.columns)
    # =-=-=-=-=-=-=-=-=-=-=-=-=-=-==-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=
    # set the respective conditions.
    rejected_max = x['QTY'] > x['AUTO ORDER'].apply(lambda x: math.ceil(x * 1.5))
    rejected_min = x['QTY'] < x['AUTO ORDER'] * 0.25
    # =-=-=-=-=-=-=-=-=-=-=-=-=-=-==-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=
    approved = x['QTY'] <= x['AUTO ORDER'].apply(lambda x: math.ceil(x * 1.5))
    approved_max = x['QTY'] == x['AUTO ORDER'].apply(lambda x: math.ceil(x * 1.5))
    approved_addition = x['AUTO ORDER'].apply(lambda x: x == 0) & x['QTY'].apply(lambda x: x <= 2 & x > 0)
    # =-=-=-=-=-=-=-=-=-=-=-=-=-=-==-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=
    # color code the data rows.
    df1 = df1.mask(rejected_max, rejected_max_style)
    df1 = df1.mask(approved, approved_style)
    df1 = df1.mask(rejected_min, rejected_min_style) 
    df1 = df1.mask(approved_max, approved_max_style)
    df1 = df1.mask(approved_addition, approved_addition_style)

    return df1

# =-=-=-=-=-=-=-=-=-=-=-=-=-=-==-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=
# =-=-=-=-=-=-=-=-=-=-=-=-=-=-==-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=
# END OF LOGIC SCRIPT, BELOW ARE SOME HELPFUL LINKS FOR ISSUES THAT FACED ME.
# =-=-=-=-=-=-=-=-=-=-=-=-=-=-==-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=
# =-=-=-=-=-=-=-=-=-=-=-=-=-=-==-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=

"""
pyinstaller -y -F -w -i "C:/Users/adam 25/Documents/Adam Tools/ico.ico" --add-data "C:/Users/adam 25/AppData/Local/Programs/Python/Python38-32/Lib/site-packages/pandas/io/formats/templates/html.tpl";"."  "C:/Users/adam 25/Documents/Adam Tools/eval.py"
"""

# https://stackoverflow.com/questions/31259673/unable-to-include-jinja2-template-to-pyinstaller-distribution
